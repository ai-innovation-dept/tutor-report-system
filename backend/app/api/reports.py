# === Phase 4: 指導報告書 CRUD START ===
import csv
import io
from collections import defaultdict
from datetime import date
from urllib.parse import quote
from uuid import UUID
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy import func, select
from sqlalchemy.orm import Session, selectinload

from app.database import get_db
from app.deps import get_current_user, get_report_for_user
from app.models import Assignment, ChatMessage, ChatRead, LessonReport, ReportEvent, ReportStatus, User
from app.schemas import ReportCreate, ReportOut, ReportPatch

STATUS_RANK = {
    ReportStatus.draft.value: 0,
    ReportStatus.returned_to_tutor.value: 0,
    ReportStatus.awaiting_parent_approval.value: 1,
    ReportStatus.parent_approved.value: 2,
    ReportStatus.submitted_to_admin.value: 3,
    ReportStatus.received.value: 4,
    ReportStatus.re_reviewed.value: 5,
    ReportStatus.admin_approved.value: 6,
}

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _current_month() -> str:
    return date.today().strftime("%Y-%m")


def _report_out(db: Session, report: LessonReport, user: User) -> ReportOut:
    last = db.scalar(select(ReportEvent.action).where(ReportEvent.report_id == report.id).order_by(ReportEvent.created_at.desc()).limit(1))
    last_return_event = db.execute(
        select(ReportEvent.comment, ReportEvent.created_at)
        .where(
            ReportEvent.report_id == report.id,
            ReportEvent.comment.is_not(None),
            ReportEvent.action.in_(["parent_return", "return_from_receiver", "return_from_reviewer", "return_from_master"]),
        )
        .order_by(ReportEvent.created_at.desc())
        .limit(1)
    ).first()
    unread = db.scalar(
        select(func.count(ChatMessage.id))
        .where(ChatMessage.report_id == report.id, ChatMessage.sender_id != user.id)
        .outerjoin(ChatRead, (ChatRead.message_id == ChatMessage.id) & (ChatRead.user_id == user.id))
        .where(ChatRead.message_id.is_(None))
    ) or 0
    out = ReportOut.model_validate(report)
    out.last_event = last
    if last_return_event:
        out.last_return_comment = last_return_event[0]
        out.last_return_at = last_return_event[1]
    out.unread_count = unread
    out.student_name = report.assignment.student_name if report.assignment else None
    out.tutor_name = report.tutor.display_name if report.tutor else None
    return out


def _teaching_minutes(report: LessonReport) -> int:
    start = report.start_time.hour * 60 + report.start_time.minute
    end = report.end_time.hour * 60 + report.end_time.minute
    return max(0, end - start - (report.break_minutes or 0))


def _latest(values) -> object | None:
    filtered = [value for value in values if value is not None]
    return max(filtered) if filtered else None


def _earliest(values) -> object | None:
    filtered = [value for value in values if value is not None]
    return min(filtered) if filtered else None


def _monthly_phase(reports: list[LessonReport]) -> str:
    statuses = [report.status for report in reports]
    ranks = [STATUS_RANK.get(status, 0) for status in statuses]
    if any(status == ReportStatus.returned_to_tutor.value for status in statuses):
        return "returned"
    if statuses and all(status == ReportStatus.admin_approved.value for status in statuses):
        return "completed"
    if statuses and all(rank >= STATUS_RANK[ReportStatus.submitted_to_admin.value] for rank in ranks):
        return "submitted_to_admin"
    if statuses and all(rank >= STATUS_RANK[ReportStatus.parent_approved.value] for rank in ranks):
        return "parent_approved"
    if statuses and all(rank >= STATUS_RANK[ReportStatus.awaiting_parent_approval.value] for rank in ranks):
        return "awaiting_parent"
    return "recording"


@router.post("", response_model=ReportOut)
def create_report(payload: ReportCreate, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    if user.role != "tutor":
        raise HTTPException(status_code=403, detail="only tutors can create reports")
    assignment = db.get(Assignment, payload.assignment_id)
    if not assignment or assignment.tutor_id != user.id or not assignment.is_active:
        raise HTTPException(status_code=403, detail="assignment access denied")
    target_month = payload.lesson_date.strftime("%Y-%m")
    current_month = _current_month()
    if target_month != current_month:
        raise HTTPException(status_code=400, detail="当月分の報告書のみ作成できます")
    existing_approved = db.scalar(
        select(func.count(LessonReport.id)).where(
            LessonReport.tutor_id == user.id,
            LessonReport.assignment_id == payload.assignment_id,
            LessonReport.target_month == current_month,
            LessonReport.status == ReportStatus.admin_approved.value,
        )
    ) or 0
    if existing_approved > 0:
        raise HTTPException(status_code=409, detail="当月分はすでに最終承認済みです。追加修正が必要な場合は運営に差戻しを依頼してください")
    report = LessonReport(
        **payload.model_dump(),
        tutor_id=user.id,
        parent_id=assignment.parent_id,
        target_month=target_month,
        status=ReportStatus.draft.value,
    )
    db.add(report)
    db.add(ReportEvent(report=report, actor_id=user.id, action="create", to_status=ReportStatus.draft.value))
    db.commit()
    db.refresh(report)
    return _report_out(db, report, user)


@router.get("", response_model=list[ReportOut])
def list_reports(status: str | None = None, target_month: str | None = None, assignment_id: UUID | None = None, tutor_id: UUID | None = None, parent_id: UUID | None = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    stmt = select(LessonReport).options(selectinload(LessonReport.assignment), selectinload(LessonReport.tutor)).order_by(LessonReport.lesson_date.desc())
    if user.role == "tutor":
        stmt = stmt.where(LessonReport.tutor_id == user.id)
    elif user.role == "parent":
        stmt = stmt.where(LessonReport.parent_id == user.id)
    elif not user.role.startswith("admin_"):
        raise HTTPException(status_code=403, detail="not allowed")
    if status:
        stmt = stmt.where(LessonReport.status == status)
    if target_month:
        stmt = stmt.where(LessonReport.target_month == target_month)
    if assignment_id:
        stmt = stmt.where(LessonReport.assignment_id == assignment_id)
    if tutor_id and (user.role == "parent" or user.role.startswith("admin_")):
        stmt = stmt.where(LessonReport.tutor_id == tutor_id)
    if parent_id and user.role.startswith("admin_"):
        stmt = stmt.where(LessonReport.parent_id == parent_id)
    return [_report_out(db, row, user) for row in db.scalars(stmt).all()]


@router.get("/monthly-summary")
def monthly_summary(tutor_id: UUID | None = None, target_month: str | None = None, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    target_tutor_id = tutor_id or user.id
    if user.role == "tutor" and target_tutor_id != user.id:
        raise HTTPException(status_code=403, detail="cannot view other tutor summary")
    if user.role != "tutor" and not user.role.startswith("admin_"):
        raise HTTPException(status_code=403, detail="not allowed")

    stmt = select(LessonReport).where(LessonReport.tutor_id == target_tutor_id)
    if target_month:
        stmt = stmt.where(LessonReport.target_month == target_month)
    reports = db.scalars(stmt.order_by(LessonReport.target_month.desc(), LessonReport.lesson_date.asc(), LessonReport.start_time.asc())).all()
    grouped: dict[str, list[LessonReport]] = defaultdict(list)
    for report in reports:
        grouped[report.target_month].append(report)

    summaries = []
    for month, items in grouped.items():
        phase = _monthly_phase(items)
        report_items = [_report_out(db, report, user).model_dump(mode="json") for report in items]
        submitted_to_parent_dates = [report.submitted_to_parent_at for report in items]
        parent_approved_dates = [report.parent_approved_at for report in items]
        submitted_to_admin_dates = [report.submitted_to_admin_at for report in items]
        admin_approved_dates = [report.admin_approved_at for report in items]
        received_dates = [report.received_at for report in items]
        re_reviewed_dates = [report.re_reviewed_at for report in items]
        summaries.append(
            {
                "target_month": month,
                "total_count": len(items),
                "total_minutes": sum(_teaching_minutes(report) for report in items),
                "phase": phase,
                "submitted_to_parent_at": _latest(submitted_to_parent_dates),
                "first_submitted_to_parent_at": _earliest(submitted_to_parent_dates),
                "parent_approved_at": _latest(parent_approved_dates),
                "submitted_to_admin_at": _latest(submitted_to_admin_dates),
                "received_at": _latest(received_dates),
                "re_reviewed_at": _latest(re_reviewed_dates),
                "admin_approved_at": _latest(admin_approved_dates),
                "has_returned": any(report.status == ReportStatus.returned_to_tutor.value for report in items),
                "can_submit_to_parent": all(report.status in {ReportStatus.draft.value, ReportStatus.returned_to_tutor.value} for report in items),
                "can_submit_to_admin": all(report.status == ReportStatus.parent_approved.value for report in items),
                "is_completed": phase == "completed",
                "reports": report_items,
                "counts_by_status": {status: sum(1 for report in items if report.status == status) for status in sorted({report.status for report in items})},
            }
        )
    return summaries


@router.get("/export")
def export_reports(
    assignment_id: UUID,
    target_month: str,
    format: str = "xlsx",
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    assignment = db.get(Assignment, assignment_id)
    if not assignment:
        raise HTTPException(status_code=404, detail="assignment not found")
    if user.role == "tutor":
        if assignment.tutor_id != user.id:
            raise HTTPException(status_code=403, detail="access denied")
    elif user.role == "parent":
        if assignment.parent_id != user.id:
            raise HTTPException(status_code=403, detail="access denied")
    elif not user.role.startswith("admin_"):
        raise HTTPException(status_code=403, detail="not allowed")

    reports = db.scalars(
        select(LessonReport)
        .where(LessonReport.assignment_id == assignment_id, LessonReport.target_month == target_month)
        .order_by(LessonReport.lesson_date, LessonReport.start_time)
    ).all()
    if not reports:
        raise HTTPException(status_code=404, detail="no reports found")

    student_name = assignment.student_name
    year, month_str = target_month.split("-")
    month_label = f"{year}年{int(month_str):02d}月"
    filename_base = f"指導実績_{student_name}_{month_label}"

    _STATUS_JA = {
        "draft": "下書き",
        "awaiting_parent_approval": "保護者承認待ち",
        "parent_approved": "保護者承認済み",
        "submitted_to_admin": "運営提出済み",
        "received": "受付済み",
        "re_reviewed": "再鑑済み",
        "admin_approved": "最終承認済み",
        "returned_to_tutor": "差戻し中",
    }
    _WD = ["日", "月", "火", "水", "木", "金", "土"]
    headers = ["回数", "指導日", "開始時刻", "終了時刻", "休憩（分）", "指導時間", "科目", "指導内容", "ステータス"]

    rows = []
    total_minutes = 0
    for i, report in enumerate(reports, 1):
        minutes = _teaching_minutes(report)
        total_minutes += minutes
        h, m = divmod(minutes, 60)
        duration = f"{h}時間{m}分" if h and m else (f"{h}時間" if h else f"{m}分")
        wd = _WD[(report.lesson_date.weekday() + 1) % 7]
        rows.append([
            i,
            f"{report.lesson_date.month}月{report.lesson_date.day}日（{wd}）",
            report.start_time.strftime("%H:%M"),
            report.end_time.strftime("%H:%M"),
            report.break_minutes or 0,
            duration,
            report.subject or "",
            report.content,
            _STATUS_JA.get(report.status, report.status),
        ])

    th, tm = divmod(total_minutes, 60)
    total_label = f"合計指導時間：" + (f"{th}時間{tm}分" if th and tm else (f"{th}時間" if th else f"{tm}分"))

    if format == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        writer.writerow([total_label])
        content_bytes = ("﻿" + buf.getvalue()).encode("utf-8")
        return Response(
            content=content_bytes,
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename_base + '.csv')}"},
        )

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{int(month_str)}月指導実績"

    bold = Font(bold=True)
    gray = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold
        cell.fill = gray

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    total_row_idx = len(rows) + 2
    total_cell = ws.cell(row=total_row_idx, column=1, value=total_label)
    total_cell.font = bold
    ws.merge_cells(start_row=total_row_idx, start_column=1, end_row=total_row_idx, end_column=len(headers))

    def _col_width(value) -> float:
        return sum(2 if ord(c) > 127 else 1 for c in str(value)) if value else 0

    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_w = max((_col_width(cell.value) for row in ws.iter_rows(min_col=col_idx, max_col=col_idx) for cell in row), default=0)
        ws.column_dimensions[col_letter].width = min(max_w + 2, 60)

    buf2 = io.BytesIO()
    wb.save(buf2)
    buf2.seek(0)
    return Response(
        content=buf2.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename_base + '.xlsx')}"},
    )


@router.get("/{report_id}", response_model=ReportOut)
def get_report(report_id: UUID, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    return _report_out(db, get_report_for_user(report_id, user, db), user)


@router.patch("/{report_id}", response_model=ReportOut)
def patch_report(report_id: UUID, payload: ReportPatch, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    report = get_report_for_user(report_id, user, db)
    if user.role != "tutor" or report.tutor_id != user.id:
        raise HTTPException(status_code=403, detail="report cannot be edited")
    if report.target_month != _current_month():
        raise HTTPException(status_code=400, detail="当月分の報告書のみ作成できます")
    if report.status not in {ReportStatus.draft.value, ReportStatus.returned_to_tutor.value}:
        raise HTTPException(status_code=409, detail="only draft or returned reports can be edited")
    data = payload.model_dump(exclude_unset=True)
    for key, value in data.items():
        setattr(report, key, value)
    if report.start_time >= report.end_time:
        raise HTTPException(status_code=422, detail="start_time must be before end_time")
    if "lesson_date" in data:
        report.target_month = report.lesson_date.strftime("%Y-%m")
        if report.target_month != _current_month():
            raise HTTPException(status_code=400, detail="当月分の報告書のみ作成できます")
    db.add(ReportEvent(report_id=report.id, actor_id=user.id, action="update", from_status=report.status, to_status=report.status))
    db.commit()
    db.refresh(report)
    return _report_out(db, report, user)


@router.delete("/{report_id}")
def delete_report(report_id: UUID, db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    report = get_report_for_user(report_id, user, db)
    if user.role != "tutor" or report.tutor_id != user.id:
        raise HTTPException(status_code=403, detail="only draft reports can be deleted")
    if report.status != ReportStatus.draft.value:
        raise HTTPException(status_code=409, detail="only draft reports can be deleted")
    message_ids = db.scalars(select(ChatMessage.id).where(ChatMessage.report_id == report.id)).all()
    if message_ids:
        db.query(ChatRead).filter(ChatRead.message_id.in_(message_ids)).delete(synchronize_session=False)
    db.query(ChatMessage).filter(ChatMessage.report_id == report.id).delete(synchronize_session=False)
    db.query(ReportEvent).filter(ReportEvent.report_id == report.id).delete(synchronize_session=False)
    db.delete(report)
    db.commit()
    return {"status": "ok"}
# === Phase 4 END ===
